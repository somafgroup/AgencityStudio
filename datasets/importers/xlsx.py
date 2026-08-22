"""Read-only XLSX importer that never executes workbook formulas or macros."""

from __future__ import annotations

from collections.abc import Iterator
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .base import BaseImporter, ImporterError, TabularSource


class XlsxImporter(BaseImporter):
    importer_id = "studio.xlsx"
    schema_version = "1"

    def _open(self, handle: BinaryIO):
        handle.seek(0)
        try:
            return load_workbook(handle, read_only=True, data_only=False, keep_links=False)
        except (InvalidFileException, OSError, ValueError, KeyError) as exc:
            raise ImporterError("The XLSX workbook could not be read.") from exc

    def open_table(self, handle: BinaryIO, *, filename: str, options: dict) -> TabularSource:
        workbook = self._open(handle)
        sheet_names = list(workbook.sheetnames)
        if not sheet_names:
            workbook.close()
            raise ImporterError("The workbook does not contain any worksheets.")
        sheet_name = str(options.get("sheet") or sheet_names[0])
        if sheet_name not in sheet_names:
            workbook.close()
            raise ImporterError("The selected worksheet does not exist in this workbook.")
        has_header = bool(options.get("has_header", True))
        worksheet = workbook[sheet_name]
        raw_rows = worksheet.iter_rows()
        try:
            first_cells = next(raw_rows)
        except StopIteration as exc:
            workbook.close()
            raise ImporterError("The selected worksheet is empty.") from exc
        first = [cell.value if cell.value is not None else "" for cell in first_cells]
        if not first or not any(value != "" for value in first):
            workbook.close()
            raise ImporterError("The selected worksheet contains no usable columns.")
        if has_header:
            source_headers = [str(value) if value is not None else "" for value in first]
            headers = [value if value.strip() else f"Column {index}" for index, value in enumerate(source_headers, 1)]
        else:
            source_headers = ["" for _ in first]
            headers = [f"Column {index}" for index in range(1, len(first) + 1)]
        metadata = {
            "used_options": {"sheet": sheet_name, "has_header": has_header},
            "source_has_header": has_header,
            "formula_cell_count": sum(1 for cell in first_cells if cell.data_type == "f"),
        }

        def rows() -> Iterator[list[object]]:
            try:
                if not has_header:
                    yield first
                for cells in raw_rows:
                    metadata["formula_cell_count"] += sum(1 for cell in cells if cell.data_type == "f")
                    yield [cell.value if cell.value is not None else "" for cell in cells]
            finally:
                workbook.close()

        return TabularSource(
            headers=headers,
            source_headers=source_headers,
            rows=rows(),
            detected_options={"sheet": sheet_name, "available_sheets": sheet_names, "has_header": True},
            metadata=metadata,
        )

    def read_page(
        self,
        handle: BinaryIO,
        *,
        filename: str,
        options: dict,
        offset: int,
        limit: int,
    ) -> tuple[list[str], list[list[object]]]:
        table = self.open_table(handle, filename=filename, options=options)
        page: list[list[object]] = []
        max_width = len(table.headers)
        for row_index, row in enumerate(table.rows):
            if row_index < offset:
                continue
            if len(page) >= limit:
                break
            values = list(row)
            page.append(values)
            max_width = max(max_width, len(values))
        if hasattr(table.rows, "close"):
            table.rows.close()
        headers = list(table.headers)
        headers.extend(f"Column {index}" for index in range(len(headers) + 1, max_width + 1))
        return headers, page
